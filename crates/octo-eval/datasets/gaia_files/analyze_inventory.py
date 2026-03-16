import openpyxl

# Read the Excel file
file_path = '/Users/sujiangwen/sandbox/LLM/speechless.ai/Autonomous-Agents/octo-sandbox/crates/octo-eval/datasets/gaia_files/32102e3e-d12a-4209-9163-7b3a104efe5d.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Print headers to understand the structure
print("Headers:", [cell.value for cell in ws[1]])
print("Number of rows:", ws.max_row)
print("\nFirst few rows of data:")

# Print all data to see the structure
for row_idx, row in enumerate(ws.iter_rows(min_val=1, max_row=min(10, ws.max_row)), 1):
    row_data = [cell.value for cell in row]
    if row_idx <= 5:  # Just show first 5 rows
        print(f"Row {row_idx}: {row_data}")